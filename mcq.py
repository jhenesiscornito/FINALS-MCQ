from tkinter import Tk, Button, Label, Entry, PhotoImage, Frame, Radiobutton, StringVar
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
import random

bg_color = "#003f5c"  # Dark Blue
fg_color = "#ffffff"  # White
button_color = "#7a5195"

def read_excel(file_path):
    try:
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active
        return sheet
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None

def center_window(window, width=500, height=400):
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    center_x = int((screen_width - width) / 2)
    center_y = int((screen_height - height) / 2)
    window.geometry(f"{width}x{height}+{center_x}+{center_y}")

def login_gui():
    username = "admin"
    password = "123123"

    login = Tk()
    login.title("Login")
    center_window(login, 500, 400)
    login.configure(bg=bg_color)

    Label(login, text="Username", bg=bg_color, fg=fg_color).place(relx=0.5, rely=0.2, anchor='center')
    username_entry = Entry(login, bg=bg_color, fg=fg_color)
    username_entry.place(relx=0.5, rely=0.3, anchor='center')

    Label(login, text="Password", bg=bg_color, fg=fg_color).place(relx=0.5, rely=0.4, anchor='center')
    password_entry = Entry(login, show="*", bg=bg_color, fg=fg_color)
    password_entry.place(relx=0.5, rely=0.5, anchor='center')

    error_label = Label(login, bg=bg_color, fg=fg_color)
    error_label.place(relx=0.5, rely=0.6, anchor='center')

    attempts = [0]

    def check_login(event=None):
        entered_username = username_entry.get()
        entered_password = password_entry.get()

        if not entered_username or not entered_password:
            error_label.config(text="Username and Password are required.")
            return

        if entered_username == username and entered_password == password:
            login.destroy()
            create_gui()
        else:
            attempts[0] += 1
            remaining_attempts = 3 - attempts[0]
            if attempts[0] < 3:
                error_label.config(
                    text=f"Incorrect Username or Password. {remaining_attempts} attempts remaining. Please try again!")
            else:
                error_label.config(text="Check with the Department Head for Account Details. Closing in 3 seconds...")
                countdown(3)

    def countdown(time_left):
        if time_left > 0:
            login.after(1000, countdown, time_left - 1)
            error_label.config(
                text=f"Check with the Department Head for Account Details. Closing in {time_left} seconds...")
        else:
            login.destroy()

    login_button = Button(login, text="Login", command=check_login, bg=button_color, fg=fg_color)
    login_button.place(relx=0.5, rely=0.7, anchor='center')

    login.bind('<Return>', check_login)

    login.mainloop()

def create_gui():
    root = Tk()
    root.title("Teacher's Tool: MCQ Randomizer")
    root.geometry("1000x700")
    root.configure(bg=bg_color)

    center_window(root, 1000, 700)

    file_path = [None]

    def select_file():
        file_path[0] = askopenfilename(filetypes=[("Excel files", "*.xlsx")])

    def upload_file():
        time_limit = time_limit_entry.get()
        if not time_limit:
            error_label.config(text="Time limit is required.")
            return

        if not file_path[0]:
            error_label.config(text="No file selected.")
            return

        ws = read_excel(file_path[0])
        if ws:
            root.destroy()
            exam_gui(ws, is_cell_bold, int(time_limit))

    frame = Frame(root, bg=bg_color)
    frame.pack()

    instructions = Label(frame, text="EXCEL FILE MUST BE IN THIS FORMAT\n\n"
                                     "1. All Questions Must be in Column A.\n"
                                     "2. All Choices from A to D must be in column B to E.\n"
                                     "NOTE: The entire question and its choices must be on row 1.\n\n"
                                     "3. Cells with correct answers must be in BOLD Letters.\n"
                                     "Here's an example layout:\n", bg=bg_color, fg=fg_color)
    instructions.pack(pady=20)

    guide_image = PhotoImage(file="guide.PNG")
    image_label = Label(frame, image=guide_image, bg=bg_color)
    image_label.image = guide_image
    image_label.pack()

    Label(frame, text="Time Limit (in minutes):", bg=bg_color, fg=fg_color).pack()

    validate_command = root.register(lambda input: input.isdigit() or input == "")
    time_limit_entry = Entry(frame, validate="key", validatecommand=(validate_command, "%P"), bg=bg_color, fg=fg_color)
    time_limit_entry.pack()

    error_label = Label(frame, text="", bg=bg_color, fg=fg_color)
    error_label.pack()

    select_button = Button(frame, text="Select Excel File", command=select_file, bg=button_color, fg=fg_color)
    select_button.pack(pady=10)

    upload_button = Button(frame, text="Upload", command=upload_file, bg=button_color, fg=fg_color)
    upload_button.pack(pady=10)

    root.mainloop()

def is_cell_bold(ws, row, col):
    cell = ws.cell(row=row, column=col)
    return cell.font.bold

def exam_gui(ws, is_cell_bold, time_limit):
    exam = Tk()
    exam.title("Student's Exam")
    exam.geometry("1000x700")

    center_window(exam, 1000, 700)

    frame = Frame(exam)
    frame.pack()

    selected_choice = StringVar(value="")  # Initialize with an empty string
    error_label = None

    questions = []
    options = []
    row_indices = []

    for idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        if row[0]:
            questions.append(row[0])
            options.append(row[1:5])
            row_indices.append(idx)

    combined = list(zip(questions, options, row_indices))
    random.shuffle(combined)
    questions, options, row_indices = zip(*combined)

    correct_answers = 0
    total_questions = len(questions)
    current_question_index = [0]

    # Convert time limit to seconds and create a label to display it
    time_left = [time_limit * 60]
    timer_label = Label(exam, text=f"Time left: {time_left[0]} seconds")
    timer_label.pack()

    def update_timer():
        if current_question_index[0] >= total_questions:  # Check if all questions have been answered
            return  # If they have, return without scheduling the next call to update_timer

        time_left[0] -= 1
        hours, remainder = divmod(time_left[0], 3600)
        minutes, seconds = divmod(remainder, 60)
        timer_label.config(text=f"Time left: {hours} hours, {minutes} minutes, {seconds} seconds")
        if time_left[0] > 0:
            # Schedule the next call to update_timer
            exam.after(1000, update_timer)

    def check_answers():
        nonlocal correct_answers, error_label
        user_answer = selected_choice.get()

        # Check if an answer is selected
        if not user_answer:
            if error_label is None:
                error_label = Label(frame, text="Please select an answer before submitting.")
                error_label.pack()
            return
        else:
            if error_label is not None:
                error_label.destroy()
                error_label = None

        correct_answer = None
        for j, option in enumerate(options[current_question_index[0]]):
            if is_cell_bold(ws, row_indices[current_question_index[0]], j + 2):
                correct_answer = chr(65 + j)
                break

        if user_answer == correct_answer:
            correct_answers += 1

        current_question_index[0] += 1
        if current_question_index[0] >= total_questions:
            for widget in frame.winfo_children():
                widget.destroy()

            score_label = Label(frame, text=f"Final Score: {correct_answers} / {total_questions}")
            score_label.pack()

            # Hide the timer label
            timer_label.pack_forget()

            # Display the remaining time
            hours, remainder = divmod(time_left[0], 3600)
            minutes, seconds = divmod(remainder, 60)
            remaining_time_label = Label(frame,
                                         text=f"Remaining time: {hours} hours, {minutes} minutes, {seconds} seconds")
            remaining_time_label.pack()

            submit_button.pack_forget()
        else:
            display_question()

    def display_question():
        for widget in frame.winfo_children():
            widget.destroy()

        question = questions[current_question_index[0]]
        question_label = Label(frame, text=question)
        question_label.pack()

        selected_choice.set(None)  # Reset the selected choice
        submit_button.config(state="disabled")  # Disable the submit button

        choices = options[current_question_index[0]]
        for idx, choice in enumerate(choices):
            choice_text = str(choice).strip()
            choice_button = Radiobutton(frame, text=choice_text, variable=selected_choice, value=chr(65 + idx),
                                        command=enable_submit)
            choice_button.pack()

    def enable_submit():
        submit_button.config(state="normal")  # Enable the submit button when a choice is selected

    submit_button = Button(exam, text="Submit", command=check_answers,
                           state="disabled")  # Initially disable the submit button
    submit_button.pack()

    # Start the timer
    update_timer()

    display_question()

    exam.mainloop()

if __name__ == "__main__":
    login_gui()